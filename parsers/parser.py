import csv
import openpyxl
import datetime
from typing import Protocol, List, Dict
from enum import Enum, auto
from dataclasses import dataclass

from logger_config import configure_logger
from parsers.georgia_mapping import GEORGIA_OPERATOR_MAPPING

logger = configure_logger(__name__)


class AvailableCountry(Enum):
    # Georgia = auto()
    # Latvia = auto()
    Kazakhstan = auto()
    Belarus = auto()


@dataclass
class ParseResult:
    hlr3_records: List[Dict[str, str]]
    hlr_records: List[Dict[str, str]]


class MnpParser(Protocol):

    def parse(self, in_file: str) -> ParseResult:
        pass


class GeorgiaMnpParser:

    def parse(self, in_file: str) -> ParseResult:
        logger.info(f"starting parsing Georgia mnp file")
        parse_result = ParseResult(hlr3_records=[], hlr_records=[])
        with open(in_file) as f:
            csv_reader = csv.reader(f, delimiter=';')
            next(csv_reader)
            for row in csv_reader:
                # record type is MNP(10) and number type is Mobile(2)
                if row[0] == '10' and row[2] == '2':
                    try:
                        mccmnc = GEORGIA_OPERATOR_MAPPING[int(row[5])]
                    except KeyError:
                        logger.warning(f'cant get mccmnc from record {row}')
                    parse_result.hlr3_records.append(
                        {
                            'dnis': row[3],
                            'mccmnc': mccmnc,
                            'active_from': int(datetime.datetime.strptime(row[9], '%Y-%m-%d %H:%M:%S').timestamp()),
                            'ownerID': None,
                            'providerResponseCode': None,
                        })
                    parse_result.hlr_records.append(
                        {
                            'dnis': row[3],
                            'mccmnc': mccmnc,
                        })
            return parse_result


class LatviaMnpParser:

    def parse(self, in_file: str) -> ParseResult:
        pass


class BelarusMnpParser:

    def parse(self, in_file: str) -> ParseResult:
        logger.info(f"starting parsing Belarus mnp file")
        parse_result = ParseResult(hlr3_records=[], hlr_records=[])
        work_book = openpyxl.load_workbook(in_file)
        sheet = work_book['Sheet1']

        for row in sheet.iter_rows(1, sheet.max_row):
            mnc_cell, msisdn_cell, port_date_cell = row
            parse_result.hlr_records.append(
                {
                    'dnis': msisdn_cell.value,
                    'mccmnc': f'2570{mnc_cell.value}',
                }
            )
            try:
                parse_result.hlr3_records.append(
                    {
                        'dnis': msisdn_cell.value,
                        'mccmnc': f'2570{mnc_cell.value}',
                        'active_from': int(
                            datetime.datetime.strptime(port_date_cell.value, '%d.%m.%Y %H:%M:%S').timestamp()),
                        'ownerID': None,
                        'providerResponseCode': None,
                    }
                )
            except:
                logger.exception(f"An error occurred while parsing record {mnc_cell.value, msisdn_cell.value, port_date_cell.value}", exc_info=True)

        return parse_result


class KazakhstanMnpParser:

    def parse(self, in_file: str) -> ParseResult:
        hlr3_records = []
        hlr_records = []
        parse_result = ParseResult(hlr3_records=hlr3_records, hlr_records=hlr_records)

        with open(in_file, 'r') as f:
            csv_reader = csv.DictReader(
                f,
                delimiter=',',
                fieldnames=('Number', 'OwnerId', 'MNC', 'Route', 'PortDate', 'RowCount')
            )
            next(csv_reader)
            for row in csv_reader:
                hlr_records = {
                    'dnis': row['Number'],
                    'mccmnc': f"4010{row['MNC']}"}
                hlr3_records = {
                    'dnis': row['Number'],
                    'mccmnc': f"4010{row['MNC']}",
                    'active_from': int(datetime.datetime.fromisoformat(row['PortDate']).timestamp()),
                    'ownerID': row['Route'],
                    'providerResponseCode': None,
                }
                parse_result.hlr_records.append(hlr_records)
                parse_result.hlr3_records.append(hlr3_records)
        return parse_result


def get_parser(country: AvailableCountry) -> MnpParser:
    try:
        match country:
            # case country.Latvia:
            #     return LatviaMnpParser()
            case country.Belarus:
                return BelarusMnpParser()
            case country.Kazakhstan:
                return KazakhstanMnpParser()
            # case country.Georgia:
            #     return GeorgiaMnpParser()
            case _:
                pass
    except AttributeError:
        logger.error('Unknown country')
